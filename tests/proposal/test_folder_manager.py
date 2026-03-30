"""
Tests for proposal/folder_manager.py and proposal/schedule_generator.py.
"""

import tempfile
from pathlib import Path
from datetime import datetime, timedelta

import pytest

from proposal.folder_manager import (
    create_proposal_folders,
    folder_summary,
    get_folder,
    place_file,
    write_readme,
)
from proposal.schedule_generator import (
    export_schedule_xlsx,
    extract_color_team_schedule,
    generate_schedule,
    schedule_to_text,
    setup_schedule,
)
from proposal.models import Proposal


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def proposal():
    return Proposal(
        id="test-001",
        title="Network Security Platform",
        solicitation_number="FA8612-26-R-0001",
        agency="AFRL",
        naics_code="541512",
        estimated_value=3_000_000,
        proposal_due_date="2026-09-01",
        capture_manager="Alice",
        proposal_manager="Bob",
    )


@pytest.fixture
def no_solicitation_proposal():
    return Proposal(id="no-sol-id", title="Unknown Opportunity")


# ---------------------------------------------------------------------------
# folder_manager: create_proposal_folders
# ---------------------------------------------------------------------------

def test_creates_all_standard_folders(proposal):
    """Expected use: all default folders are created under the solicitation root."""
    with tempfile.TemporaryDirectory() as tmp:
        base = Path(tmp)
        result = create_proposal_folders(proposal, base_dir=base)
        assert "root" in result
        assert result["root"].exists()
        # spot-check key folders
        assert (result["root"] / "00_rfp").exists()
        assert (result["root"] / "03_working" / "vol_1_technical").exists()
        assert (result["root"] / "_admin").exists()


def test_idempotent_create(proposal):
    """Edge case: calling create_proposal_folders twice doesn't raise or lose files."""
    with tempfile.TemporaryDirectory() as tmp:
        base = Path(tmp)
        create_proposal_folders(proposal, base_dir=base)
        create_proposal_folders(proposal, base_dir=base)  # second call
        assert (base / "FA8612-26-R-0001" / "00_rfp").exists()


def test_folder_root_uses_solicitation_number(proposal):
    """Expected use: root folder is named after the solicitation number."""
    with tempfile.TemporaryDirectory() as tmp:
        result = create_proposal_folders(proposal, base_dir=Path(tmp))
        assert result["root"].name == "FA8612-26-R-0001"


def test_folder_root_falls_back_to_id(no_solicitation_proposal):
    """Edge case: no solicitation number → falls back to proposal ID prefix."""
    with tempfile.TemporaryDirectory() as tmp:
        result = create_proposal_folders(no_solicitation_proposal, base_dir=Path(tmp))
        assert result["root"].name == "no-sol-id"[:8]


# ---------------------------------------------------------------------------
# folder_manager: folder_summary
# ---------------------------------------------------------------------------

def test_folder_summary_all_exist_after_create(proposal):
    """Expected use: summary shows 'exists' for all folders after creation."""
    with tempfile.TemporaryDirectory() as tmp:
        base = Path(tmp)
        create_proposal_folders(proposal, base_dir=base)
        summary = folder_summary(proposal, base_dir=base)
        missing = [k for k, v in summary.items() if v == "missing"]
        assert missing == []


def test_folder_summary_missing_before_create(proposal):
    """Failure case: summary shows 'missing' before folders are created."""
    with tempfile.TemporaryDirectory() as tmp:
        summary = folder_summary(proposal, base_dir=Path(tmp))
        assert summary["root"] == "missing"


# ---------------------------------------------------------------------------
# folder_manager: place_file / get_folder
# ---------------------------------------------------------------------------

def test_place_file_returns_correct_path(proposal):
    """Expected use: place_file returns path and creates the folder."""
    with tempfile.TemporaryDirectory() as tmp:
        target = place_file(proposal, "02_bid_no_bid", "test.pptx", base_dir=Path(tmp))
        assert target.name == "test.pptx"
        assert target.parent.exists()


def test_get_folder_returns_path(proposal):
    """Expected use: get_folder returns a Path without creating it."""
    with tempfile.TemporaryDirectory() as tmp:
        p = get_folder(proposal, "00_rfp", base_dir=Path(tmp))
        assert isinstance(p, Path)
        assert "00_rfp" in str(p)


# ---------------------------------------------------------------------------
# folder_manager: write_readme
# ---------------------------------------------------------------------------

def test_write_readme_creates_file(proposal):
    """Expected use: README.txt written to proposal root."""
    with tempfile.TemporaryDirectory() as tmp:
        readme = write_readme(proposal, base_dir=Path(tmp))
        assert readme.exists()
        content = readme.read_text()
        assert proposal.solicitation_number in content
        assert proposal.title in content
        assert proposal.capture_manager in content


# ---------------------------------------------------------------------------
# schedule_generator: generate_schedule
# ---------------------------------------------------------------------------

def test_generate_schedule_count(proposal):
    """Expected use: schedule produces 12 default milestones."""
    milestones = generate_schedule(proposal)
    assert len(milestones) == 12


def test_schedule_sorted_chronologically(proposal):
    """Expected use: milestones are in ascending date order."""
    milestones = generate_schedule(proposal)
    dates = [m.date for m in milestones]
    assert dates == sorted(dates)


def test_submission_day_is_due_date(proposal):
    """Expected use: last milestone date equals the proposal due date."""
    milestones = generate_schedule(proposal)
    last = milestones[-1]
    assert last.date == "2026-09-01"
    assert last.days_before_due == 0


def test_kickoff_is_45_days_before(proposal):
    """Expected use: kickoff is 45 calendar days before due date."""
    milestones = generate_schedule(proposal)
    kickoff = next(m for m in milestones if "Kickoff" in m.name)
    due = datetime.fromisoformat(proposal.proposal_due_date)
    kickoff_date = datetime.fromisoformat(kickoff.date)
    assert (due - kickoff_date).days == 45


def test_generate_schedule_no_due_date_raises():
    """Failure case: missing due date raises ValueError."""
    p = Proposal(id="x", title="No Date")
    with pytest.raises(ValueError, match="proposal_due_date not set"):
        generate_schedule(p)


# ---------------------------------------------------------------------------
# schedule_generator: extract_color_team_schedule
# ---------------------------------------------------------------------------

def test_color_team_dates_extracted(proposal):
    """Expected use: Pink, Red, Gold team dates are populated from schedule."""
    milestones = generate_schedule(proposal)
    cts = extract_color_team_schedule(milestones)
    assert cts.pink_team is not None
    assert cts.red_team is not None
    assert cts.gold_team is not None


def test_color_team_order(proposal):
    """Edge case: Pink < Red < Gold in chronological order."""
    milestones = generate_schedule(proposal)
    cts = extract_color_team_schedule(milestones)
    assert cts.pink_team < cts.red_team < cts.gold_team  # type: ignore[operator]


# ---------------------------------------------------------------------------
# schedule_generator: schedule_to_text
# ---------------------------------------------------------------------------

def test_schedule_text_contains_solicitation(proposal):
    """Expected use: text output includes solicitation number."""
    milestones = generate_schedule(proposal)
    text = schedule_to_text(milestones, proposal)
    assert proposal.solicitation_number in text
    assert "Kickoff" in text
    assert "Submission" in text


# ---------------------------------------------------------------------------
# schedule_generator: export_schedule_xlsx
# ---------------------------------------------------------------------------

def test_xlsx_export_creates_file(proposal):
    """Expected use: export_schedule_xlsx creates a readable .xlsx file."""
    milestones = generate_schedule(proposal)
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "schedule.xlsx"
        result = export_schedule_xlsx(milestones, proposal, out)
        assert result.exists()
        assert result.suffix == ".xlsx"
        assert result.stat().st_size > 5_000


def test_xlsx_submission_row_present(proposal):
    """Edge case: submission day row exists in the xlsx data."""
    import openpyxl
    milestones = generate_schedule(proposal)
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "schedule.xlsx"
        export_schedule_xlsx(milestones, proposal, out)
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        cell_values = [ws.cell(row=r, column=3).value for r in range(1, ws.max_row + 1)]  # type: ignore[union-attr]
        assert "Submission Day" in cell_values


# ---------------------------------------------------------------------------
# schedule_generator: setup_schedule (integration)
# ---------------------------------------------------------------------------

def test_setup_schedule_returns_all_keys(proposal):
    """Expected use: setup_schedule result has all expected keys."""
    with tempfile.TemporaryDirectory() as tmp:
        result = setup_schedule(proposal, admin_folder=Path(tmp))
        assert "milestones" in result
        assert "color_teams" in result
        assert "xlsx_path" in result
        assert "text_summary" in result
        assert result["xlsx_path"].exists()
