"""
Tests for proposal/cost_estimator.py

Covers:
    - Expected use: rate calculation, XLSX generation, estimate build
    - Edge cases: zero hours, single period, empty sub-costs
    - Failure cases: bad rate file, missing fields
"""

import json
from pathlib import Path
from unittest.mock import patch
import os

import pytest

from proposal.cost_estimator import (
    CostEstimate,
    LaborCategory,
    LaborLine,
    OdcLine,
    build_estimate_from_dict,
    export_cost_xlsx,
    export_rate_file,
    load_rate_file,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def clean_rates(monkeypatch):
    """Set known, predictable rates for all tests."""
    monkeypatch.setenv("FRINGE_RATE", "0.30")
    monkeypatch.setenv("OVERHEAD_RATE", "0.15")
    monkeypatch.setenv("GA_RATE", "0.10")
    monkeypatch.setenv("FEE_RATE", "0.08")


@pytest.fixture
def pm_lcat(clean_rates):
    return LaborCategory(
        name="Program Manager",
        level="Senior",
        base_hourly=100.0,
    )


@pytest.fixture
def dev_lcat(clean_rates):
    return LaborCategory(
        name="Software Developer",
        level="Mid",
        base_hourly=80.0,
    )


@pytest.fixture
def simple_estimate(clean_rates, pm_lcat, dev_lcat):
    pm_line = LaborLine(
        clin="0001", slin="0001AA", lcat=pm_lcat,
        period_hours={0: 960, 1: 960},
    )
    dev_line = LaborLine(
        clin="0001", slin="0001AB", lcat=dev_lcat,
        period_hours={0: 1920, 1: 1920},
    )
    odc = OdcLine(clin="9001", description="Travel", quantity=2, unit="LOT", unit_cost=5000.0, period=0)
    return CostEstimate(
        solicitation_number="FA8612-25-R-0001",
        proposal_title="Test Proposal",
        contract_type="ffp",
        periods=2,
        labor_lines=[pm_line, dev_line],
        odc_lines=[odc],
        subcontractor_costs={0: 50000.0, 1: 50000.0},
    )


# ---------------------------------------------------------------------------
# LaborCategory
# ---------------------------------------------------------------------------

class TestLaborCategory:
    def test_fully_burdened_rate_formula(self, clean_rates):
        lcat = LaborCategory(name="Test", base_hourly=100.0)
        # 100 × 1.30 × 1.15 × 1.10 × 1.08
        expected = 100.0 * 1.30 * 1.15 * 1.10 * 1.08
        assert abs(lcat.fully_burdened_rate() - expected) < 0.01

    def test_instance_override_takes_precedence(self, clean_rates):
        lcat = LaborCategory(name="Test", base_hourly=100.0, fringe_rate=0.50)
        # Use 0.50 fringe instead of env 0.30
        expected = 100.0 * 1.50 * 1.15 * 1.10 * 1.08
        assert abs(lcat.fully_burdened_rate() - expected) < 0.01

    def test_zero_base_rate(self, clean_rates):
        lcat = LaborCategory(name="Test", base_hourly=0.0)
        assert lcat.fully_burdened_rate() == 0.0

    def test_rate_breakdown_keys(self, clean_rates):
        lcat = LaborCategory(name="Test", base_hourly=100.0)
        breakdown = lcat.rate_breakdown()
        assert set(breakdown.keys()) == {"base_hourly", "after_fringe", "after_overhead", "after_ga", "fully_burdened"}

    def test_rate_breakdown_progression_ascending(self, clean_rates):
        lcat = LaborCategory(name="Test", base_hourly=100.0)
        b = lcat.rate_breakdown()
        assert b["base_hourly"] < b["after_fringe"] < b["after_overhead"] < b["after_ga"] < b["fully_burdened"]


# ---------------------------------------------------------------------------
# LaborLine
# ---------------------------------------------------------------------------

class TestLaborLine:
    def test_total_hours_sums_all_periods(self, pm_lcat, clean_rates):
        line = LaborLine("0001", "0001AA", pm_lcat, period_hours={0: 1000, 1: 500, 2: 250})
        assert line.total_hours() == 1750.0

    def test_total_cost_equals_hours_times_rate(self, pm_lcat, clean_rates):
        line = LaborLine("0001", "0001AA", pm_lcat, period_hours={0: 1920})
        assert abs(line.total_cost() - 1920 * pm_lcat.fully_burdened_rate()) < 0.01

    def test_zero_hours_period_zero_cost(self, pm_lcat, clean_rates):
        line = LaborLine("0001", "0001AA", pm_lcat, period_hours={0: 0})
        assert line.total_cost() == 0.0

    def test_cost_by_period_keys_match_period_hours(self, pm_lcat, clean_rates):
        line = LaborLine("0001", "0001AA", pm_lcat, period_hours={0: 1000, 1: 800})
        costs = line.cost_by_period()
        assert set(costs.keys()) == {0, 1}

    def test_empty_period_hours(self, pm_lcat, clean_rates):
        line = LaborLine("0001", "0001AA", pm_lcat, period_hours={})
        assert line.total_hours() == 0.0
        assert line.total_cost() == 0.0


# ---------------------------------------------------------------------------
# OdcLine
# ---------------------------------------------------------------------------

class TestOdcLine:
    def test_total_cost(self):
        odc = OdcLine("9001", "Travel", 3, "LOT", 2500.0)
        assert odc.total_cost() == 7500.0

    def test_zero_quantity(self):
        odc = OdcLine("9001", "Travel", 0, "LOT", 2500.0)
        assert odc.total_cost() == 0.0

    def test_fractional_quantity(self):
        odc = OdcLine("9001", "Licenses", 2.5, "YR", 10000.0)
        assert odc.total_cost() == 25000.0


# ---------------------------------------------------------------------------
# CostEstimate
# ---------------------------------------------------------------------------

class TestCostEstimate:
    def test_total_labor_cost(self, simple_estimate):
        expected = sum(line.total_cost() for line in simple_estimate.labor_lines)
        assert abs(simple_estimate.total_labor_cost() - expected) < 0.01

    def test_total_odc_cost(self, simple_estimate):
        assert simple_estimate.total_odc_cost() == 10000.0  # 2 × $5000

    def test_total_subcontractor_cost(self, simple_estimate):
        assert simple_estimate.total_subcontractor_cost() == 100000.0

    def test_grand_total_sums_components(self, simple_estimate):
        expected = (
            simple_estimate.total_labor_cost()
            + simple_estimate.total_odc_cost()
            + simple_estimate.total_subcontractor_cost()
        )
        assert abs(simple_estimate.grand_total() - expected) < 0.01

    def test_period_totals_has_correct_periods(self, simple_estimate):
        totals = simple_estimate.period_totals()
        assert 0 in totals
        assert 1 in totals

    def test_summary_dict_keys(self, simple_estimate):
        summary = simple_estimate.summary_dict()
        expected_keys = {
            "solicitation_number", "proposal_title", "contract_type", "periods",
            "total_labor", "total_odc", "total_subcontractor", "grand_total",
            "period_totals", "labor_category_count", "total_labor_hours",
        }
        assert expected_keys <= set(summary.keys())

    def test_empty_estimate_grand_total_zero(self, clean_rates):
        est = CostEstimate(solicitation_number="X", proposal_title="Y")
        assert est.grand_total() == 0.0


# ---------------------------------------------------------------------------
# load_rate_file / export_rate_file
# ---------------------------------------------------------------------------

class TestRateFileRoundtrip:
    def test_export_and_reload(self, tmp_path, clean_rates):
        lcats = [
            LaborCategory("PM", "Senior", 100.0),
            LaborCategory("Dev", "Mid", 80.0),
        ]
        path = tmp_path / "rates.json"
        export_rate_file(lcats, path)
        reloaded = load_rate_file(path)
        assert len(reloaded) == 2
        assert reloaded[0].name == "PM"
        assert reloaded[1].base_hourly == 80.0

    def test_load_missing_file_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            load_rate_file(tmp_path / "nope.json")

    def test_load_non_array_raises(self, tmp_path):
        p = tmp_path / "bad.json"
        p.write_text('{"name": "PM"}')
        with pytest.raises(ValueError):
            load_rate_file(p)

    def test_rate_overrides_preserved_in_export(self, tmp_path, clean_rates):
        lcat = LaborCategory("Dev", "Mid", 80.0, fringe_rate=0.40)
        path = tmp_path / "rates.json"
        export_rate_file([lcat], path)
        reloaded = load_rate_file(path)
        assert reloaded[0].fringe_rate == 0.40


# ---------------------------------------------------------------------------
# export_cost_xlsx
# ---------------------------------------------------------------------------

class TestExportCostXlsx:
    def test_creates_xlsx_file(self, simple_estimate, tmp_path, monkeypatch, clean_rates):
        monkeypatch.setenv("OUTPUTS_DIR", str(tmp_path))
        path = export_cost_xlsx(simple_estimate)
        assert path.is_file()
        assert path.suffix == ".xlsx"

    def test_filename_contains_solicitation(self, simple_estimate, tmp_path, monkeypatch, clean_rates):
        monkeypatch.setenv("OUTPUTS_DIR", str(tmp_path))
        path = export_cost_xlsx(simple_estimate)
        assert "FA8612-25-R-0001" in path.name

    def test_explicit_output_path(self, simple_estimate, tmp_path, clean_rates):
        out = tmp_path / "test_cost.xlsx"
        path = export_cost_xlsx(simple_estimate, output_path=out)
        assert path == out
        assert out.is_file()

    def test_workbook_has_four_sheets(self, simple_estimate, tmp_path, monkeypatch, clean_rates):
        from openpyxl import load_workbook
        monkeypatch.setenv("OUTPUTS_DIR", str(tmp_path))
        path = export_cost_xlsx(simple_estimate)
        wb = load_workbook(path)
        assert set(wb.sheetnames) == {"Cover", "Labor", "ODC", "Summary"}


# ---------------------------------------------------------------------------
# build_estimate_from_dict
# ---------------------------------------------------------------------------

class TestBuildEstimateFromDict:
    def test_builds_from_complete_dict(self, clean_rates):
        data = {
            "solicitation_number": "TEST-001",
            "proposal_title": "My Proposal",
            "contract_type": "t_and_m",
            "periods": 2,
            "labor_lines": [
                {
                    "clin": "0001", "slin": "0001AA",
                    "lcat_name": "Dev", "lcat_level": "Senior",
                    "base_hourly": 90.0,
                    "period_hours": {"0": 1920, "1": 1920},
                }
            ],
            "odc_lines": [
                {"clin": "9001", "description": "Travel", "quantity": 1, "unit": "LOT", "unit_cost": 3000.0}
            ],
            "subcontractor_costs": {"0": 20000.0},
        }
        est = build_estimate_from_dict(data)
        assert est.solicitation_number == "TEST-001"
        assert len(est.labor_lines) == 1
        assert est.labor_lines[0].period_hours[0] == 1920.0
        assert est.total_odc_cost() == 3000.0
        assert est.total_subcontractor_cost() == 20000.0

    def test_empty_dict_returns_empty_estimate(self, clean_rates):
        est = build_estimate_from_dict({})
        assert est.grand_total() == 0.0
        assert est.labor_lines == []

    def test_grand_total_gt_zero_with_labor(self, clean_rates):
        data = {
            "solicitation_number": "X",
            "proposal_title": "Y",
            "labor_lines": [
                {"lcat_name": "Dev", "base_hourly": 100.0, "period_hours": {"0": 100}}
            ],
        }
        est = build_estimate_from_dict(data)
        assert est.grand_total() > 0.0
