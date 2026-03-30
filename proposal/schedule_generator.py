"""
Proposal Schedule Generator.

Generates a working-backwards milestone schedule from the proposal due date.
Offsets are config-driven via PROPOSAL_SCHEDULE_FILE env var or defaults.

Also produces an Excel schedule file (.xlsx) via openpyxl for placement in
the _admin/ proposal folder.
"""

from __future__ import annotations

import json
import logging
import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

from proposal.models import ColorTeamSchedule, Proposal

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Milestone definitions
# ---------------------------------------------------------------------------

@dataclass
class Milestone:
    """
    A single proposal schedule milestone.

    Attributes:
        name: Milestone name.
        days_before_due: Negative integer — days before the proposal due date.
        owner: Role responsible (e.g., "Capture Manager", "Proposal Manager").
        notes: Optional context note.
        date: Computed ISO date string (set by generate_schedule).
    """
    name: str
    days_before_due: int       # negative = before due date
    owner: str = ""
    notes: str = ""
    date: str = ""             # populated by generate_schedule()


# Reason: Offset table is the single source of truth — no dates hardcoded anywhere
_DEFAULT_MILESTONES: List[Dict] = [
    {"name": "Kickoff Meeting",           "days_before_due": -45, "owner": "Proposal Manager",  "notes": "Launch proposal effort; all hands"},
    {"name": "Storyboard Complete",       "days_before_due": -35, "owner": "Volume Leads",       "notes": "All outlines/storyboards approved by PM"},
    {"name": "Pink Team Draft Due",       "days_before_due": -28, "owner": "Volume Leads",       "notes": "First complete draft submitted for review"},
    {"name": "Pink Team Review",          "days_before_due": -25, "owner": "Proposal Manager",   "notes": "Internal review session; Red comments distributed"},
    {"name": "Pink Team Comments Due",    "days_before_due": -23, "owner": "Volume Leads",       "notes": "All Pink Team comments incorporated"},
    {"name": "Red Team Draft Due",        "days_before_due": -18, "owner": "Volume Leads",       "notes": "Revised draft for Red Team"},
    {"name": "Red Team Review",           "days_before_due": -15, "owner": "Proposal Manager",   "notes": "Red Team session; senior reviewer panel"},
    {"name": "Red Team Comments Due",     "days_before_due": -13, "owner": "Volume Leads",       "notes": "All Red Team comments incorporated"},
    {"name": "Gold Team Draft Due",       "days_before_due": -7,  "owner": "Proposal Manager",   "notes": "Final content submitted for executive review"},
    {"name": "Gold Team Review",          "days_before_due": -5,  "owner": "Capture Manager",    "notes": "Executive review; price-to-win final check"},
    {"name": "Production Complete",       "days_before_due": -2,  "owner": "Proposal Manager",   "notes": "Final editing, formatting, compliance check"},
    {"name": "Submission Day",            "days_before_due":  0,  "owner": "Proposal Manager",   "notes": "Submit by 4:00 PM local time"},
]


def _load_milestones() -> List[Dict]:
    """
    Load milestone definitions from JSON file if configured, else use defaults.

    Returns:
        List[Dict]: Milestone definition dicts.
    """
    cfg_file = os.getenv("PROPOSAL_SCHEDULE_FILE", "")
    if cfg_file and Path(cfg_file).exists():
        with open(cfg_file) as f:
            data = json.load(f)
        return data.get("milestones", _DEFAULT_MILESTONES)
    return _DEFAULT_MILESTONES


# ---------------------------------------------------------------------------
# Schedule generation
# ---------------------------------------------------------------------------

def generate_schedule(proposal: Proposal) -> List[Milestone]:
    """
    Generate a milestone schedule working backwards from the proposal due date.

    Args:
        proposal: Proposal with proposal_due_date set (ISO format).

    Returns:
        List[Milestone]: Milestones in chronological order with dates populated.

    Raises:
        ValueError: If proposal_due_date is not set.
    """
    if not proposal.proposal_due_date:
        raise ValueError(f"proposal_due_date not set for proposal {proposal.id}")

    due_date = datetime.fromisoformat(proposal.proposal_due_date[:10])
    milestone_defs = _load_milestones()

    milestones: List[Milestone] = []
    for defn in milestone_defs:
        offset = defn["days_before_due"]
        target_date = due_date + timedelta(days=offset)
        m = Milestone(
            name=defn["name"],
            days_before_due=offset,
            owner=defn.get("owner", ""),
            notes=defn.get("notes", ""),
            date=target_date.strftime("%Y-%m-%d"),
        )
        milestones.append(m)

    # Sort chronologically
    milestones.sort(key=lambda m: m.date)
    return milestones


def extract_color_team_schedule(milestones: List[Milestone]) -> ColorTeamSchedule:
    """
    Extract color team review dates from the full milestone list.

    Args:
        milestones: Output of generate_schedule().

    Returns:
        ColorTeamSchedule: Dates for Pink, Red, Gold team reviews.
    """
    dates: Dict[str, Optional[str]] = {
        "pink_team": None,
        "red_team": None,
        "gold_team": None,
        "orals": None,
    }
    for m in milestones:
        name_lower = m.name.lower()
        if "pink team review" in name_lower:
            dates["pink_team"] = m.date
        elif "red team review" in name_lower:
            dates["red_team"] = m.date
        elif "gold team review" in name_lower:
            dates["gold_team"] = m.date
        elif "oral" in name_lower:
            dates["orals"] = m.date
    return ColorTeamSchedule(**dates)


def schedule_to_text(milestones: List[Milestone], proposal: Proposal) -> str:
    """
    Format the milestone schedule as a plain-text table.

    Args:
        milestones: Output of generate_schedule().
        proposal: Proposal for header context.

    Returns:
        str: Formatted schedule text.
    """
    lines = [
        f"Proposal Schedule: {proposal.title}",
        f"Solicitation: {proposal.solicitation_number or 'TBD'}",
        f"Due Date: {proposal.proposal_due_date or 'TBD'}",
        "",
        f"{'Date':<12}  {'Days':<6}  {'Milestone':<35}  {'Owner':<22}  Notes",
        "-" * 110,
    ]
    for m in milestones:
        days_label = f"D{m.days_before_due}" if m.days_before_due != 0 else "D-Day"
        lines.append(
            f"{m.date:<12}  {days_label:<6}  {m.name:<35}  {m.owner:<22}  {m.notes}"
        )
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_schedule_xlsx(
    milestones: List[Milestone],
    proposal: Proposal,
    output_path: Path,
) -> Path:
    """
    Export the milestone schedule to an Excel .xlsx file.

    Args:
        milestones: Output of generate_schedule().
        proposal: Proposal for header context.
        output_path: Full path to write the .xlsx file.

    Returns:
        Path: Absolute path to the written file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required for Excel export — run: uv add openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proposal Schedule"  # type: ignore[assignment]

    # Reason: Dark blue header row matches company branding defaults
    header_fill = PatternFill("solid", fgColor="002060")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill = PatternFill("solid", fgColor="D9E1F2")

    # Title rows
    ws.merge_cells("A1:E1")  # type: ignore[attr-defined]
    ws["A1"] = f"Proposal Schedule — {proposal.title}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Solicitation: {proposal.solicitation_number or 'TBD'}"
    ws["A3"] = f"Due Date: {proposal.proposal_due_date or 'TBD'}"
    ws["A4"] = f"Agency: {proposal.agency or 'TBD'}"
    ws.append([])  # blank row

    # Header row
    headers = ["Date", "Days Before Due", "Milestone", "Owner", "Notes"]
    ws.append(headers)
    header_row = ws.max_row  # type: ignore[attr-defined]
    for col in range(1, 6):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, m in enumerate(milestones):
        days_label = f"D{m.days_before_due}" if m.days_before_due != 0 else "D-Day"
        ws.append([m.date, days_label, m.name, m.owner, m.notes])
        data_row = ws.max_row  # type: ignore[attr-defined]
        # Alternating row fill
        if i % 2 == 1:
            for col in range(1, 6):
                ws.cell(row=data_row, column=col).fill = alt_fill
        # Highlight submission day
        if m.days_before_due == 0:
            submission_fill = PatternFill("solid", fgColor="00B050")
            submission_font = Font(bold=True, color="FFFFFF")
            for col in range(1, 6):
                ws.cell(row=data_row, column=col).fill = submission_fill
                ws.cell(row=data_row, column=col).font = submission_font

    # Column widths
    col_widths = {"A": 13, "B": 16, "C": 36, "D": 24, "E": 50}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width  # type: ignore[attr-defined]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Saved proposal schedule XLSX: %s", output_path)
    return output_path.resolve()


# ---------------------------------------------------------------------------
# Convenience: full setup helper used by proposal-setup skill
# ---------------------------------------------------------------------------

def setup_schedule(
    proposal: Proposal,
    admin_folder: Optional[Path] = None,
) -> Dict:
    """
    Generate schedule, export XLSX, return structured result.

    Args:
        proposal: Proposal model with proposal_due_date set.
        admin_folder: Path to the _admin/ folder for the proposal.
            Defaults to outputs/proposal/{solicitation}/_admin/.

    Returns:
        Dict with keys: milestones, color_teams, xlsx_path, text_summary.
    """
    milestones = generate_schedule(proposal)
    color_teams = extract_color_team_schedule(milestones)
    text = schedule_to_text(milestones, proposal)

    sol_id = (proposal.solicitation_number or proposal.id[:8]).replace("/", "-")
    date_str = datetime.now().strftime("%Y-%m-%d")

    if admin_folder is None:
        admin_folder = Path("outputs") / "proposal" / sol_id / "_admin"

    xlsx_path = export_schedule_xlsx(
        milestones,
        proposal,
        admin_folder / f"{sol_id}_proposal_schedule_{date_str}.xlsx",
    )

    return {
        "milestones":    milestones,
        "color_teams":   color_teams,
        "xlsx_path":     xlsx_path,
        "text_summary":  text,
    }
