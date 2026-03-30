"""
Cost Estimator — Labor/wrap rate calculator for GovCon proposals.

Builds fully-burdened labor rates, generates Section B price schedules,
and rolls up CLIN-level costs by period.

Configuration via environment variables:
    FRINGE_RATE        — decimal fringe benefit rate (default: 0.30)
    OVERHEAD_RATE      — decimal overhead rate (default: 0.15)
    GA_RATE            — decimal G&A rate (default: 0.10)
    FEE_RATE           — decimal profit/fee rate (default: 0.08)
    CONTRACT_TYPE      — ffp | t_and_m | labor_hour (default: ffp)
    OUTPUTS_DIR        — base output path (default: outputs/proposal)
"""

import json
import logging
import os
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

def _rate(env_key: str, default: float) -> float:
    """Load a decimal rate from env, falling back to default."""
    try:
        return float(os.getenv(env_key, str(default)))
    except ValueError:
        logger.warning("Invalid %s, using default %.2f", env_key, default)
        return default


def _outputs_dir() -> Path:
    return Path(os.getenv("OUTPUTS_DIR", Path(__file__).parent.parent / "outputs" / "proposal"))


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class LaborCategory:
    """
    A single Labor Category (LCAT) definition.

    Attributes:
        name: LCAT name (e.g., 'Program Manager').
        level: Seniority level ('Junior', 'Mid', 'Senior', 'Principal').
        base_hourly: Base hourly rate (before fringe/overhead/G&A/fee).
        fringe_rate: Decimal fringe benefit rate (overrides env if set).
        overhead_rate: Decimal overhead rate (overrides env if set).
        ga_rate: Decimal G&A rate (overrides env if set).
        fee_rate: Decimal profit/fee rate (overrides env if set).
    """
    name: str
    level: str = "Mid"
    base_hourly: float = 0.0
    fringe_rate: Optional[float] = None
    overhead_rate: Optional[float] = None
    ga_rate: Optional[float] = None
    fee_rate: Optional[float] = None

    @property
    def effective_fringe(self) -> float:
        return self.fringe_rate if self.fringe_rate is not None else _rate("FRINGE_RATE", 0.30)

    @property
    def effective_overhead(self) -> float:
        return self.overhead_rate if self.overhead_rate is not None else _rate("OVERHEAD_RATE", 0.15)

    @property
    def effective_ga(self) -> float:
        return self.ga_rate if self.ga_rate is not None else _rate("GA_RATE", 0.10)

    @property
    def effective_fee(self) -> float:
        return self.fee_rate if self.fee_rate is not None else _rate("FEE_RATE", 0.08)

    def fully_burdened_rate(self) -> float:
        """
        Calculate fully-burdened hourly rate.

        Formula: base × (1 + fringe) × (1 + overhead) × (1 + G&A) × (1 + fee)

        Returns:
            float: Fully-burdened hourly rate in dollars.
        """
        return (
            self.base_hourly
            * (1 + self.effective_fringe)
            * (1 + self.effective_overhead)
            * (1 + self.effective_ga)
            * (1 + self.effective_fee)
        )

    def rate_breakdown(self) -> Dict[str, float]:
        """
        Return step-by-step rate buildup for cost narrative.

        Returns:
            Dict[str, float]: Ordered rate components.
        """
        base = self.base_hourly
        after_fringe    = base * (1 + self.effective_fringe)
        after_overhead  = after_fringe * (1 + self.effective_overhead)
        after_ga        = after_overhead * (1 + self.effective_ga)
        fully_burdened  = after_ga * (1 + self.effective_fee)
        return {
            "base_hourly":      round(base, 2),
            "after_fringe":     round(after_fringe, 2),
            "after_overhead":   round(after_overhead, 2),
            "after_ga":         round(after_ga, 2),
            "fully_burdened":   round(fully_burdened, 2),
        }


@dataclass
class LaborLine:
    """
    One labor line item in a cost estimate (LCAT × period hours).

    Attributes:
        clin: Contract Line Item Number (e.g., '0001').
        slin: Sub-Line Item Number (e.g., '0001AA').
        lcat: Labor category definition.
        period_hours: Hours per period of performance (period index → hours).
        task: Task or work package name.
    """
    clin: str
    slin: str
    lcat: LaborCategory
    period_hours: Dict[int, float] = field(default_factory=dict)  # period_idx → hours
    task: str = ""

    def total_hours(self) -> float:
        """Total hours across all periods."""
        return sum(self.period_hours.values())

    def total_cost(self) -> float:
        """Total fully-burdened cost across all periods."""
        return self.total_hours() * self.lcat.fully_burdened_rate()

    def cost_by_period(self) -> Dict[int, float]:
        """Fully-burdened cost by period."""
        rate = self.lcat.fully_burdened_rate()
        return {p: h * rate for p, h in self.period_hours.items()}


@dataclass
class OdcLine:
    """
    Other Direct Cost (ODC) line item.

    Attributes:
        clin: Contract Line Item Number.
        description: Description of the ODC.
        quantity: Number of units.
        unit: Unit of measure (e.g., 'EA', 'LOT', 'HR').
        unit_cost: Cost per unit.
        period: Period of performance index (0 = base, 1+ = options).
    """
    clin: str
    description: str
    quantity: float
    unit: str
    unit_cost: float
    period: int = 0

    def total_cost(self) -> float:
        """Total cost for this ODC line."""
        return self.quantity * self.unit_cost


@dataclass
class CostEstimate:
    """
    Full cost estimate for a proposal.

    Attributes:
        solicitation_number: RFP solicitation number.
        proposal_title: Title of the proposal.
        contract_type: Contract type (ffp, t_and_m, labor_hour).
        periods: Number of periods (base + option years).
        labor_lines: All labor line items.
        odc_lines: All other direct cost lines.
        subcontractor_costs: Sub-tier costs per period.
        notes: Free-form notes for cost narrative.
    """
    solicitation_number: str
    proposal_title: str
    contract_type: str = "ffp"
    periods: int = 1
    labor_lines: List[LaborLine] = field(default_factory=list)
    odc_lines: List[OdcLine] = field(default_factory=list)
    subcontractor_costs: Dict[int, float] = field(default_factory=dict)  # period → amount
    notes: str = ""

    def total_labor_cost(self) -> float:
        """Total fully-burdened labor cost across all periods."""
        return sum(line.total_cost() for line in self.labor_lines)

    def total_odc_cost(self) -> float:
        """Total ODC cost."""
        return sum(odc.total_cost() for odc in self.odc_lines)

    def total_subcontractor_cost(self) -> float:
        """Total subcontractor costs."""
        return sum(self.subcontractor_costs.values())

    def grand_total(self) -> float:
        """Total proposal price."""
        return self.total_labor_cost() + self.total_odc_cost() + self.total_subcontractor_cost()

    def period_totals(self) -> Dict[int, float]:
        """
        Calculate total cost by period (base + each option year).

        Returns:
            Dict[int, float]: period_index → total cost.
        """
        totals: Dict[int, float] = {}
        for line in self.labor_lines:
            for period, cost in line.cost_by_period().items():
                totals[period] = totals.get(period, 0.0) + cost
        for odc in self.odc_lines:
            totals[odc.period] = totals.get(odc.period, 0.0) + odc.total_cost()
        for period, amt in self.subcontractor_costs.items():
            totals[period] = totals.get(period, 0.0) + amt
        return totals

    def summary_dict(self) -> Dict:
        """
        Return a JSON-serializable summary for reporting.

        Returns:
            Dict: Key cost figures and breakdowns.
        """
        return {
            "solicitation_number": self.solicitation_number,
            "proposal_title": self.proposal_title,
            "contract_type": self.contract_type,
            "periods": self.periods,
            "total_labor": round(self.total_labor_cost(), 2),
            "total_odc": round(self.total_odc_cost(), 2),
            "total_subcontractor": round(self.total_subcontractor_cost(), 2),
            "grand_total": round(self.grand_total(), 2),
            "period_totals": {str(k): round(v, 2) for k, v in self.period_totals().items()},
            "labor_category_count": len({line.lcat.name for line in self.labor_lines}),
            "total_labor_hours": round(sum(line.total_hours() for line in self.labor_lines), 1),
        }


# ---------------------------------------------------------------------------
# Rate file import / export
# ---------------------------------------------------------------------------

def load_rate_file(path: Path) -> List[LaborCategory]:
    """
    Load LCAT definitions from a JSON rate file.

    Expected format: list of objects with name, level, base_hourly, and
    optional rate override fields.

    Args:
        path: Path to the JSON rate file.

    Returns:
        List[LaborCategory]: Parsed labor categories.

    Raises:
        FileNotFoundError: If path does not exist.
        ValueError: If format is invalid.
    """
    if not path.is_file():
        raise FileNotFoundError(f"Rate file not found: {path}")

    with open(path) as f:
        raw = json.load(f)

    if not isinstance(raw, list):
        raise ValueError(f"Rate file must be a JSON array, got {type(raw).__name__}")

    lcats = []
    for item in raw:
        lcats.append(LaborCategory(
            name=str(item.get("name", "Unknown")),
            level=str(item.get("level", "Mid")),
            base_hourly=float(item.get("base_hourly", 0.0)),
            fringe_rate=item.get("fringe_rate"),
            overhead_rate=item.get("overhead_rate"),
            ga_rate=item.get("ga_rate"),
            fee_rate=item.get("fee_rate"),
        ))
    logger.info("Loaded %d LCATs from %s", len(lcats), path)
    return lcats


def export_rate_file(lcats: List[LaborCategory], path: Path) -> None:
    """
    Export LCAT definitions to a JSON rate file.

    Args:
        lcats: Labor category definitions to export.
        path: Output file path.
    """
    data = []
    for lcat in lcats:
        entry = {
            "name": lcat.name,
            "level": lcat.level,
            "base_hourly": lcat.base_hourly,
        }
        if lcat.fringe_rate is not None:
            entry["fringe_rate"] = lcat.fringe_rate
        if lcat.overhead_rate is not None:
            entry["overhead_rate"] = lcat.overhead_rate
        if lcat.ga_rate is not None:
            entry["ga_rate"] = lcat.ga_rate
        if lcat.fee_rate is not None:
            entry["fee_rate"] = lcat.fee_rate
        data.append(entry)

    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        json.dump(data, f, indent=2)
    logger.info("Exported %d LCATs to %s", len(lcats), path)


# ---------------------------------------------------------------------------
# XLSX Section B generation
# ---------------------------------------------------------------------------

def export_cost_xlsx(estimate: CostEstimate, output_path: Optional[Path] = None) -> Path:
    """
    Generate a Section B price schedule XLSX workbook.

    Sheets:
        Cover       — metadata and totals
        Labor       — CLIN/SLIN labor lines with period columns
        ODC         — other direct costs
        Summary     — period-by-period and grand total roll-up

    Args:
        estimate: Fully populated CostEstimate.
        output_path: Where to write the file. Auto-generated if None.

    Returns:
        Path: Path to the written XLSX file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    _navy   = "002060"
    _blue   = "0070C0"
    _gray   = "F2F2F2"
    _green  = "E2EFDA"
    _white  = "FFFFFF"

    def _hdr(ws, row, col, val, fill_color=_navy, bold=True, wrap=True):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = Font(name="Calibri", bold=bold, color=_white if fill_color == _navy or fill_color == _blue else "000000", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

    def _val(ws, row, col, val, fmt=None, bold=False):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name="Calibri", bold=bold, size=10)
        cell.alignment = Alignment(vertical="center")
        if fmt:
            cell.number_format = fmt
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        return cell

    # ---- Cover sheet ----
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "SECTION B — PRICE SCHEDULE"
    cover["A1"].font = Font(name="Calibri", bold=True, size=16, color=_navy)
    cover.merge_cells("A1:D1")

    meta = [
        ("Solicitation Number", estimate.solicitation_number),
        ("Proposal Title",      estimate.proposal_title),
        ("Contract Type",       estimate.contract_type.upper()),
        ("Periods of Performance", str(estimate.periods)),
        ("Total Labor",         estimate.total_labor_cost()),
        ("Total ODC",           estimate.total_odc_cost()),
        ("Total Subcontractor", estimate.total_subcontractor_cost()),
        ("GRAND TOTAL",         estimate.grand_total()),
    ]
    for i, (label, val) in enumerate(meta, start=3):
        bold = label == "GRAND TOTAL"
        cover.cell(i, 1, label).font = Font(name="Calibri", bold=True, size=10)
        cell = cover.cell(i, 2, val)
        cell.font = Font(name="Calibri", bold=bold, size=10)
        if isinstance(val, float):
            cell.number_format = '"$"#,##0.00'
    cover.column_dimensions["A"].width = 28
    cover.column_dimensions["B"].width = 40

    # ---- Labor sheet ----
    labor = wb.create_sheet("Labor")
    period_labels = [f"Period {p}" for p in range(estimate.periods)]
    period_cost_labels = [f"Period {p} Cost" for p in range(estimate.periods)]
    base_headers = ["CLIN", "SLIN", "Labor Category", "Level", "Fully-Burdened Rate"]
    all_headers = base_headers + period_labels + period_cost_labels + ["Total Hours", "Total Cost"]

    for col, h in enumerate(all_headers, start=1):
        _hdr(labor, 1, col, h)

    for row_idx, line in enumerate(estimate.labor_lines, start=2):
        rate = line.lcat.fully_burdened_rate()
        _val(labor, row_idx, 1, line.clin)
        _val(labor, row_idx, 2, line.slin)
        _val(labor, row_idx, 3, line.lcat.name)
        _val(labor, row_idx, 4, line.lcat.level)
        _val(labor, row_idx, 5, round(rate, 2), fmt='"$"#,##0.00')

        for p_idx, p_label in enumerate(period_labels):
            col = 6 + p_idx
            hours = line.period_hours.get(p_idx, 0.0)
            _val(labor, row_idx, col, hours)

        for p_idx in range(estimate.periods):
            col = 6 + estimate.periods + p_idx
            hours = line.period_hours.get(p_idx, 0.0)
            _val(labor, row_idx, col, round(hours * rate, 2), fmt='"$"#,##0.00')

        total_col_h = 6 + estimate.periods * 2
        total_col_c = total_col_h + 1
        _val(labor, row_idx, total_col_h, round(line.total_hours(), 1), bold=True)
        _val(labor, row_idx, total_col_c, round(line.total_cost(), 2), fmt='"$"#,##0.00', bold=True)

    for i, letter in enumerate("ABCDE"):
        labor.column_dimensions[letter].width = [8, 10, 28, 10, 18][i]

    # ---- ODC sheet ----
    odc = wb.create_sheet("ODC")
    odc_headers = ["CLIN", "Description", "Quantity", "Unit", "Unit Cost", "Total Cost", "Period"]
    for col, h in enumerate(odc_headers, start=1):
        _hdr(odc, 1, col, h, fill_color=_blue)
    for row_idx, line in enumerate(estimate.odc_lines, start=2):
        _val(odc, row_idx, 1, line.clin)
        _val(odc, row_idx, 2, line.description)
        _val(odc, row_idx, 3, line.quantity)
        _val(odc, row_idx, 4, line.unit)
        _val(odc, row_idx, 5, round(line.unit_cost, 2), fmt='"$"#,##0.00')
        _val(odc, row_idx, 6, round(line.total_cost(), 2), fmt='"$"#,##0.00')
        _val(odc, row_idx, 7, f"Period {line.period}")
    odc.column_dimensions["B"].width = 40

    # ---- Summary sheet ----
    summary = wb.create_sheet("Summary")
    summary["A1"] = "COST SUMMARY BY PERIOD"
    summary["A1"].font = Font(name="Calibri", bold=True, size=14, color=_navy)
    summary.merge_cells("A1:C1")

    _hdr(summary, 2, 1, "Period", fill_color=_navy)
    _hdr(summary, 2, 2, "Labor",  fill_color=_navy)
    _hdr(summary, 2, 3, "ODC",    fill_color=_navy)
    _hdr(summary, 2, 4, "Subcontractor", fill_color=_navy)
    _hdr(summary, 2, 5, "Period Total", fill_color=_navy)

    period_totals = estimate.period_totals()
    for row_idx, p_idx in enumerate(range(estimate.periods), start=3):
        labor_cost = sum(line.cost_by_period().get(p_idx, 0.0) for line in estimate.labor_lines)
        odc_cost   = sum(odc.total_cost() for odc in estimate.odc_lines if odc.period == p_idx)
        sub_cost   = estimate.subcontractor_costs.get(p_idx, 0.0)
        total      = labor_cost + odc_cost + sub_cost
        _val(summary, row_idx, 1, f"Period {p_idx} {'(Base)' if p_idx == 0 else f'(Option {p_idx})'}")
        _val(summary, row_idx, 2, round(labor_cost, 2), fmt='"$"#,##0.00')
        _val(summary, row_idx, 3, round(odc_cost, 2),   fmt='"$"#,##0.00')
        _val(summary, row_idx, 4, round(sub_cost, 2),   fmt='"$"#,##0.00')
        _val(summary, row_idx, 5, round(total, 2),       fmt='"$"#,##0.00', bold=True)

    grand_row = 3 + estimate.periods
    _val(summary, grand_row, 1, "GRAND TOTAL", bold=True)
    _val(summary, grand_row, 5, round(estimate.grand_total(), 2), fmt='"$"#,##0.00', bold=True)

    for col in ["A", "B", "C", "D", "E"]:
        summary.column_dimensions[col].width = 28

    # ---- Write file ----
    if output_path is None:
        safe_sol = estimate.solicitation_number.replace("/", "-").replace(" ", "_")
        today = date.today().strftime("%Y-%m-%d")
        out_dir = _outputs_dir() / safe_sol
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = out_dir / f"{safe_sol}_cost_volume_{today}.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Saved cost volume: %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# Convenience builders
# ---------------------------------------------------------------------------

def build_estimate_from_dict(data: Dict) -> CostEstimate:
    """
    Build a CostEstimate from a plain dictionary (e.g., from API or JSON).

    Expected keys:
        solicitation_number (str)
        proposal_title (str)
        contract_type (str, optional)
        periods (int, optional)
        labor_lines (list of dicts)
        odc_lines (list of dicts, optional)
        subcontractor_costs (dict str→float, optional)

    Each labor_line dict:
        clin, slin, lcat_name, lcat_level, base_hourly, period_hours (dict)

    Args:
        data: Input dictionary.

    Returns:
        CostEstimate: Populated estimate object.
    """
    labor_lines = []
    for ll in data.get("labor_lines", []):
        lcat = LaborCategory(
            name=str(ll.get("lcat_name", "Unknown")),
            level=str(ll.get("lcat_level", "Mid")),
            base_hourly=float(ll.get("base_hourly", 0.0)),
        )
        period_hours = {int(k): float(v) for k, v in ll.get("period_hours", {}).items()}
        labor_lines.append(LaborLine(
            clin=str(ll.get("clin", "0001")),
            slin=str(ll.get("slin", "0001AA")),
            lcat=lcat,
            period_hours=period_hours,
            task=str(ll.get("task", "")),
        ))

    odc_lines = []
    for ol in data.get("odc_lines", []):
        odc_lines.append(OdcLine(
            clin=str(ol.get("clin", "9001")),
            description=str(ol.get("description", "")),
            quantity=float(ol.get("quantity", 1)),
            unit=str(ol.get("unit", "LOT")),
            unit_cost=float(ol.get("unit_cost", 0.0)),
            period=int(ol.get("period", 0)),
        ))

    sub_costs = {int(k): float(v) for k, v in data.get("subcontractor_costs", {}).items()}

    return CostEstimate(
        solicitation_number=str(data.get("solicitation_number", "")),
        proposal_title=str(data.get("proposal_title", "")),
        contract_type=str(data.get("contract_type", "ffp")),
        periods=int(data.get("periods", 1)),
        labor_lines=labor_lines,
        odc_lines=odc_lines,
        subcontractor_costs=sub_costs,
        notes=str(data.get("notes", "")),
    )
